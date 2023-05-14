import psycopg2 as pg
import re
import openpyxl as opxl


filename='Employee Dat.xlsx'
split_name = re.split(r'\.|\s', filename)
# print(split_name)
dbname=split_name[0].lower()
table_name=split_name[0]+split_name[1]

try:
    with open(filename,'rt') as file:
        print("File exists and connection is establishing.....")

#Establishing connection...............
    connection=pg.connect(user="postgres",password="Manoj@009", host="localhost",port=5432)
    cursors=connection.cursor()
    connection.rollback()
    connection.autocommit=True
#terminating all the active connectivity ----------

    cursors.execute("SELECT pg_terminate_backend(pg_stat_activity.pid) "
                "FROM pg_stat_activity "
                "WHERE pg_stat_activity.datname = 'employee' "
                "AND pid <> pg_backend_pid();")
    
#Drop and creating database---------

    cursors.execute("DROP DATABASE IF EXISTS "+dbname)
    cursors.execute("CREATE DATABASE "+dbname)
    print("database named {} is created successfully".format(dbname))
    cursors.close()
    connection.close()
    connection=pg.connect(database=dbname,user="postgres",password="Manoj@009", host="localhost",port=5432)
    cursors=connection.cursor()

#Creating table .............

    commands="""
        CREATE TABLE if not exists {} (
        S_No int,
        First_Name varchar(20),
        Last_Name varchar(10),
        Gender char(15),Country varchar(20),
        Age int,
        Date varchar(10),
        Employee_Id varchar(20)
        )
            """.format(table_name
                       )
    cursors.execute(commands)
    connection.commit()
    print("Table {} is created.....".format(table_name))
    file_path=filename

    #Excel file is  loaded------------

    wb=opxl.load_workbook(file_path)
    sheet=wb.active
    cursors.execute("Select * FROM " +table_name+" LIMIT 0")
    column_names = [desc[0] for desc in cursors.description]
    # print(column_names)
    current_header=[cell.value for cell in sheet[1]]
    # print(current_header) 
    if current_header != column_names:
        for index, column_name in enumerate(column_names):
            cell = sheet.cell(row=1, column=index+1)
            cell.value = column_name
        wb.save(file_path)
    print("Employee Data headers are taken care of")

#data is inserted into table

    for row in sheet.iter_rows(min_row=2, values_only=True):
        insert_query = f"INSERT INTO {table_name} VALUES {tuple(row)};"
        cursors.execute(insert_query)
    connection.commit()

  

    cursors.close()
    connection.close()

    print("Automation is completed")
except FileNotFoundError as e:
    raise e
    









































      # query = f"SELECT employee_id,first_name,last_name,gender,date,country FROM {table_name} limit 10;"
    # cursors.execute(query)
    # rows = cursors.fetchall()
    # column_names = [desc[0] for desc in cursors.description]
    # print(column_names)
    # for row in rows:
    #     print(row)

    # query = f"SELECT first_name FROM {table_name} where country='France' limit 10;"
    # cursors.execute(query)
    # rows = cursors.fetchall()
    # print(rows)
    # column_names = [desc[0] for desc in cursors.description]
    # print(column_names)
    # for row in rows:
    #     print(row)